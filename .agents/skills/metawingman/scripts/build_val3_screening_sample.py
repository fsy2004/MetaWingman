"""Build the frozen VAL-3 AI screening sample (preregistered 2026-08-18).

- Gold positives: merged-corpus records whose normalized title matches a
  2022 workbook included-study title (frozen rule: lowercase, alphanumerics
  only, exact match).
- Sample = ALL gold matches + 500 seeded non-gold records (seed 20260815).
- Writes sample JSONL + manifest (hashes). Read-only on inputs.

Usage:
  python metawingman/scripts/build_val3_screening_sample.py \
    --workbook <2022 xlsx> --corpus <merged-candidates.jsonl> \
    --out-dir <dir>
"""
from __future__ import annotations

import argparse
import hashlib
import json
import random
import re
from pathlib import Path


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def norm_title(title: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (title or "").lower())


def workbook_titles(workbook: Path) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["References"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {name: i for i, name in enumerate(header) if name}
    title_col = idx["Title"]
    type_col = idx["Type of paper"]
    study_col = idx["Study ID"]
    gold = []
    for row in rows[1:]:
        if row is None or not row[study_col]:
            continue
        paper_type = str(row[type_col] or "").strip().lower()
        if paper_type not in ("evaluation", "case-control"):
            continue
        title = str(row[title_col] or "").strip()
        if title:
            gold.append({"study_id": str(row[study_col]), "title": title, "norm": norm_title(title)})
    wb.close()
    return [g["norm"] for g in gold], gold


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--corpus", type=Path, required=True)
    parser.add_argument("--out-dir", type=Path, required=True)
    parser.add_argument("--seed", type=int, default=20260815)
    parser.add_argument("--non-gold", type=int, default=500)
    args = parser.parse_args()
    try:
        gold_norms, gold_meta = workbook_titles(args.workbook)
        corpus = [json.loads(line) for line in args.corpus.read_text(encoding="utf-8-sig").splitlines() if line.strip()]
        corpus_norms = {norm_title(r.get("title") or ""): r for r in corpus}
        matches = []
        matched_norms = set()
        for g in gold_meta:
            hit = corpus_norms.get(g["norm"])
            if hit is not None and g["norm"] not in matched_norms:
                matches.append({"gold_study_id": g["study_id"], "gold_title": g["title"], **hit})
                matched_norms.add(g["norm"])
        non_gold = [r for r in corpus if norm_title(r.get("title") or "") not in matched_norms]
        rng = random.Random(args.seed)
        sample_non_gold = rng.sample(non_gold, min(args.non_gold, len(non_gold)))
        args.out_dir.mkdir(parents=True, exist_ok=False)
        sample = [
            {"role": "gold", "gold_study_id": m["gold_study_id"], "gold_title": m["gold_title"],
             "id": m.get("id"), "title": m.get("title"), "abstract": m.get("abstract") or "",
             "first_publication_date": m.get("first_publication_date")}
            for m in matches
        ] + [
            {"role": "unlabeled", "id": r.get("id"), "title": r.get("title"), "abstract": r.get("abstract") or "",
             "first_publication_date": r.get("first_publication_date")}
            for r in sample_non_gold
        ]
        sample_path = args.out_dir / "screening-sample.jsonl"
        sample_path.write_text("\n".join(json.dumps(s, ensure_ascii=False) for s in sample) + "\n", encoding="utf-8")
        manifest = {
            "schema_version": "1.0",
            "gold_titles_in_workbook": len(gold_meta),
            "gold_matched_in_corpus": len(matches),
            "gold_unmatched": len(gold_meta) - len(matches),
            "unmatched_gold_study_ids": [g["study_id"] for g in gold_meta if g["norm"] not in matched_norms],
            "non_gold_sampled": len(sample_non_gold),
            "seed": args.seed,
            "sample_total": len(sample),
            "sample_sha256": sha256_file(sample_path),
            "workbook_sha256": sha256_file(args.workbook),
            "corpus_sha256": sha256_file(args.corpus),
            "matching_rule": "lowercase alphanumerics only, exact match",
        }
        (args.out_dir / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(json.dumps(manifest, indent=2, ensure_ascii=False))
        return 0
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        print(json.dumps({"execution_state": "failed", "error": str(exc)}, indent=2))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
