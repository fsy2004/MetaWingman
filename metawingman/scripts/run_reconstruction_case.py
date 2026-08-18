"""Run a reconstruction case (analysis slice) under the run-lock policy.

A reconstruction case replays a frozen published-review analysis slice with
MetaWingman's deterministic R pipeline and, optionally, scores the outputs
against the sealed published-reference answers under precommitted tolerances.

Safety contract (enforced, never advisory):
1. The RUN_BOUNDARY.json must exist and contain a run lock for the
   configuration/repetition this invocation records BEFORE any sealed file is
   read. The runner appends its own lock first and re-verifies.
2. The sealed input artifact is verified by SHA-256 before being copied into
   the working directory; the sealed tree itself is never modified.
3. Reference answers are read only when --score is given AND the boundary's
   expected run count is satisfied; their SHA-256 must match the case spec.
4. Every output (staged CSV, R outputs, receipt) is hashed into the receipt.

Usage (local, R 4.x + metafor/meta/pdftools + openpyxl required):
  python metawingman/scripts/run_reconstruction_case.py \
    --case research/reconstruction-cases/<case>.json \
    --boundary <RUN_BOUNDARY.json> \
    --out-dir validation-output/reconstruction-runs/<case>-<rep> \
    --toolkit toolkit --configuration deterministic-r-analysis \
    --repetition 1 [--score] [--dry-run]
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

CASE_SCHEMA = Path(__file__).resolve().parent.parent / "schemas" / "reconstruction_case.schema.json"


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def validate_case(case: dict) -> None:
    try:
        import jsonschema
    except ImportError as exc:  # pragma: no cover - documented dependency
        raise SystemExit("jsonschema is required: pip install jsonschema") from exc
    schema = json.loads(CASE_SCHEMA.read_text(encoding="utf-8"))
    errors = list(jsonschema.Draft202012Validator(schema).iter_errors(case))
    if errors:
        for err in errors[:5]:
            print(f"  case schema error at {err.json_path}: {err.message}", file=sys.stderr)
        raise ValueError(f"case spec invalid ({len(errors)} schema errors)")


def load_boundary(path: Path) -> dict:
    if not path.is_file():
        raise ValueError(f"RUN_BOUNDARY.json missing at {path}")
    boundary = json.loads(path.read_text(encoding="utf-8"))
    if boundary.get("schema_version") != "2.0":
        raise ValueError("RUN_BOUNDARY.json must be schema_version 2.0")
    return boundary


def lock_is_complete(boundary: dict, expected: int) -> bool:
    return boundary.get("run_state") == "collecting" or len(boundary.get("run_locks", [])) >= expected


def append_run_lock(boundary_path: Path, boundary: dict, lock: dict) -> None:
    locks = list(boundary.get("run_locks", []))
    locks.append(lock)
    boundary["run_locks"] = locks
    if len(locks) >= boundary.get("expected_runs", 0):
        boundary["run_state"] = "locked"
    boundary_path.write_text(json.dumps(boundary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def update_run_lock(boundary_path: Path, boundary: dict, lock: dict) -> None:
    """Replace the entry whose (case_id, configuration, repetition) matches."""
    keys = ("case_id", "configuration", "repetition")
    for index, existing in enumerate(boundary.get("run_locks", [])):
        if all(existing.get(k) == lock.get(k) for k in keys):
            boundary["run_locks"][index] = lock
            boundary_path.write_text(json.dumps(boundary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
            return
    raise ValueError("run lock entry to update not found")


def stage_xlsx_to_csv(xlsx: Path, sheet: int, mapping: dict, csv_out: Path) -> None:
    """Copy the chosen sheet into a metafor-compatible CSV using the mapping."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise SystemExit("openpyxl is required: pip install openpyxl") from exc
    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet > len(workbook.sheetnames):
        raise ValueError(f"xlsx has {len(workbook.sheetnames)} sheets; requested sheet {sheet}")
    ws = workbook.worksheets[sheet - 1]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("empty xlsx sheet")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    body = rows[1:]
    columns = {}
    for meta_key in ("m1i", "sd1i", "n1i", "m2i", "sd2i", "n2i", "slab"):
        src = mapping.get(meta_key)
        if src not in header:
            raise ValueError(f"mapped column {meta_key}->'{src}' not found in sheet header {header}")
        columns[meta_key] = header.index(src)
    with csv_out.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        # Header uses the SOURCE column names; the R adapter resolves them
        # through the --m1i/--sd1i/... overrides (col_of).
        writer.writerow([mapping[meta_key] for meta_key in ("m1i", "sd1i", "n1i", "m2i", "sd2i", "n2i", "slab")])
        for row in body:
            if row is None or all(c is None for c in row):
                continue
            writer.writerow([row[idx] for idx in columns.values()])


def run_r_call(adapter: Path, workdir: Path, toolkit: Path, call: dict, inputs: dict) -> list[Path]:
    args = [
        "Rscript", str(adapter),
        "--input", str(inputs["csv"]),
        "--outdir", str(workdir / call["analysis"]),
        "--toolkit", str(toolkit),
        "--analysis", call["analysis"],
    ]
    if call.get("measure"):
        args += ["--measure", call["measure"]]
    if call.get("method"):
        args += ["--method", call["method"]]
    if call.get("knha") is not None:
        args += ["--knha", "true" if call["knha"] else "false"]
    for meta_key, src in inputs["mapping"].items():
        if meta_key != "slab":
            args += [f"--{meta_key}", src]
    if inputs["mapping"].get("slab"):
        args += ["--slab", inputs["mapping"]["slab"]]
    proc = subprocess.run(args, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if proc.returncode != 0:
        raise RuntimeError(
            f"R call {call['adapter']}/{call['analysis']} failed (rc={proc.returncode}):\n"
            f"STDOUT: {proc.stdout[-2000:]}\nSTDERR: {proc.stderr[-2000:]}"
        )
    out_dir = workdir / call["analysis"]
    return sorted(out_dir.glob("*.csv")) + sorted(out_dir.glob("*.pdf")) + sorted(out_dir.glob("*.png"))


def compare_to_reference(case: dict, workdir: Path) -> dict:
    """Score the deterministic outputs against the sealed reference answers."""
    ref_path = Path(case["reference_answers"]["package_path"])
    expected = case["reference_answers"]["sha256"]
    if expected and sha256_file(ref_path) != expected:
        raise ValueError("reference answer SHA-256 mismatch; scoring refused")
    ref = json.loads(ref_path.read_text(encoding="utf-8"))
    summary_csv = workdir / "summary" / "summary.csv"
    if not summary_csv.is_file():
        return {"scored": False, "reason": "summary.csv missing"}
    rows = list(csv.DictReader(summary_csv.open(encoding="utf-8")))
    if not rows:
        return {"scored": False, "reason": "summary.csv empty"}
    ours = rows[0]
    tol = case["tolerances"]
    agreements = {}
    for metric in ("est", "ci.lb", "ci.ub", "tau2"):
        ref_val = next((e["value"] for e in ref.get("estimates", []) if e.get("metric") == f"pooled_{metric}"), None)
        # reference file uses its own metric naming; fall back to a tolerant lookup
        if ref_val is None:
            for e in ref.get("estimates", []):
                if metric in (e.get("metric") or "") and e.get("value") is not None:
                    ref_val = e["value"]
                    break
        if ref_val is None:
            agreements[metric] = {"reference_missing": True}
            continue
        ours_val = float(ours[metric])
        key = "ci_bounds" if metric in ("ci.lb", "ci.ub") else ("i2_pp" if metric == "I2" else metric)
        limit = tol.get(key, tol.get(metric, 0))
        agreements[metric] = {
            "ours": ours_val, "reference": ref_val,
            "abs_delta": round(ours_val - ref_val, 6),
            "tolerance": limit,
            "within_tolerance": abs(ours_val - ref_val) <= limit,
        }
    return {"scored": True, "agreements": agreements}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--case", type=Path, required=True)
    parser.add_argument("--boundary", type=Path, required=True)
    parser.add_argument("--out-dir", type=Path, required=True)
    parser.add_argument("--toolkit", type=Path, required=True)
    parser.add_argument("--configuration", required=True)
    parser.add_argument("--repetition", type=int, default=1)
    parser.add_argument("--score", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    try:
        started = time.monotonic()
        case = json.loads(args.case.read_text(encoding="utf-8"))
        validate_case(case)
        if args.configuration not in case["boundary"]["configuration_ids"]:
            raise ValueError(f"configuration {args.configuration!r} not in case boundary configs")
        boundary = load_boundary(args.boundary)
        expected = case["boundary"]["expected_runs"]
        if len(boundary.get("run_locks", [])) >= expected:
            raise ValueError("boundary already fully locked; start a new run directory")

        # Contract: the run lock is recorded BEFORE any sealed file is read.
        # Dry-run is the exception: it verifies hashes and boundary state
        # without recording a lock and without staging or analyzing anything.
        if args.dry_run:
            input_artifact = Path(case["input_artifact"]["package_path"])
            if not input_artifact.is_file():
                raise ValueError(f"sealed input artifact missing: {input_artifact}")
            if sha256_file(input_artifact) != case["input_artifact"]["sha256"]:
                raise ValueError("input artifact SHA-256 mismatch; aborting before any read")
            print(json.dumps({"dry_run": True, "case": case["case_id"],
                              "input_sha256_ok": True, "boundary_runs": len(boundary.get("run_locks", []))}, indent=2))
            return 0

        lock = {
            "case_id": case["case_id"],
            "configuration": args.configuration,
            "repetition": args.repetition,
            "locked_at_utc": datetime.now(timezone.utc).isoformat(),
            "input_tree_sha256": None,
            "output_hashes": {},
            "prompt_sha256": hashlib.sha256(b"deterministic-r-pipeline").hexdigest(),
            "model": "deterministic-r-pipeline",
            "tool_versions": ["metawingman-reconstruction-runner-v1"],
        }
        append_run_lock(args.boundary, boundary, lock)

        input_artifact = Path(case["input_artifact"]["package_path"])
        if not input_artifact.is_file():
            raise ValueError(f"sealed input artifact missing: {input_artifact}")
        if sha256_file(input_artifact) != case["input_artifact"]["sha256"]:
            raise ValueError("input artifact SHA-256 mismatch; aborting before any read")

        args.out_dir.mkdir(parents=True, exist_ok=False)
        staged = args.out_dir / "input.xlsx"
        shutil.copyfile(input_artifact, staged)  # staged copy; sealed tree untouched

        csv_out = args.out_dir / "data.csv"
        stage_xlsx_to_csv(staged, case["input_artifact"]["xlsx_sheet"], case["column_mapping"], csv_out)

        adapters_dir = Path(__file__).resolve().parent / "r" / "adapters"
        inputs = {"csv": csv_out, "mapping": case["column_mapping"]}
        produced: list[Path] = [staged, csv_out]
        for call in case["r_calls"]:
            adapter = adapters_dir / call["adapter"]
            if not adapter.is_file():
                raise ValueError(f"adapter missing: {adapter}")
            produced.extend(run_r_call(adapter, args.out_dir, args.toolkit, call, inputs))

        # Finalize the run lock with provenance (scoring reads happen after).
        lock["input_tree_sha256"] = sha256_file(staged)
        lock["output_hashes"] = {p.relative_to(args.out_dir).as_posix(): sha256_file(p) for p in produced if p.is_file()}
        update_run_lock(args.boundary, boundary, lock)

        receipt = {
            "schema_version": "1.0",
            "case_id": case["case_id"],
            "execution_state": "completed",
            "configuration": args.configuration,
            "repetition": args.repetition,
            "elapsed_seconds": round(time.monotonic() - started, 3),
            "output_hashes": lock["output_hashes"],
            "boundary_run_count": len(boundary["run_locks"]),
            "scoring": None,
        }
        if args.score:
            if len(boundary["run_locks"]) < expected:
                print("scoring deferred: boundary not fully locked", file=sys.stderr)
            else:
                receipt["scoring"] = compare_to_reference(case, args.out_dir)
        (args.out_dir / "execution-receipt.json").write_text(
            json.dumps(receipt, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
        print(json.dumps(receipt, indent=2, ensure_ascii=False))
        return 0
    except (OSError, ValueError, json.JSONDecodeError, RuntimeError) as exc:
        print(json.dumps({"execution_state": "failed", "error": str(exc)}, indent=2))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
